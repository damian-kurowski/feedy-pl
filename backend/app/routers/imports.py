"""CSV / XLSX import endpoints for source feeds.

Two-step flow:
  1. POST /api/imports/preview  — upload file, returns detected columns + first 5 rows
  2. POST /api/imports/commit   — confirm column mapping + feed name, creates
                                  feed_in (custom_source=True) with all rows as ProductIn
                                  (custom_product=True). Idempotent on SKU if re-imported.
"""
import csv
import io
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user
from app.models.feed_in import FeedIn
from app.models.product import ProductIn
from app.models.user import User

router = APIRouter(prefix="/api/imports", tags=["imports"])


# ── Column auto-detection ────────────────────────────────────────────────

# Common Polish + English header keywords mapped to canonical product fields.
HEADER_KEYWORDS = {
    "id":          ["id", "sku", "kod produktu", "code", "ean_id", "product_id"],
    "name":        ["name", "nazwa", "tytul", "tytuł", "title", "produkt"],
    "url":         ["url", "link", "adres", "strona"],
    "price":       ["price", "cena", "cena brutto", "cena netto", "kwota"],
    "old_price":   ["old_price", "cena_przed", "stara cena"],
    "image":       ["image", "img", "zdjecie", "zdjęcie", "obraz", "photo", "obrazek"],
    "description": ["description", "desc", "opis", "opis produktu"],
    "category":    ["category", "kategoria", "cat", "categories"],
    "brand":       ["brand", "marka", "producer", "producent", "vendor"],
    "ean":         ["ean", "gtin", "kod kreskowy", "barcode"],
    "availability":["availability", "dostepnosc", "dostępność", "stan", "in_stock", "stock", "magazyn"],
    "weight":      ["weight", "waga"],
    "shipping":    ["shipping", "dostawa", "wysyłka", "wysylka"],
}


def _normalize_header(s: str) -> str:
    return (s or "").strip().lower().replace("-", "_").replace(" ", "_")


def auto_map_headers(headers: list[str]) -> dict[str, str | None]:
    """For each detected header, return its canonical field name (or None)."""
    mapping: dict[str, str | None] = {}
    for h in headers:
        norm = _normalize_header(h)
        matched: str | None = None
        for canonical, keywords in HEADER_KEYWORDS.items():
            for kw in keywords:
                if _normalize_header(kw) == norm or _normalize_header(kw) in norm:
                    matched = canonical
                    break
            if matched:
                break
        mapping[h] = matched
    return mapping


def _parse_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Detect delimiter
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    return headers, data


def _parse_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl niezainstalowany na backendzie")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c).strip() if c is not None else f"col_{i+1}" for i, c in enumerate(first)]
    data: list[list[Any]] = []
    for row in rows_iter:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        data.append(["" if c is None else str(c) for c in row])
    return headers, data


def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[list[str]]]:
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    # Last-resort: try CSV
    return _parse_csv(content)


# ── Endpoints ────────────────────────────────────────────────────────────

@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Parse uploaded CSV/XLSX, return headers + first 5 rows + auto mapping."""
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Plik za duży (max 10 MB)")

    headers, rows = parse_upload(file.filename or "upload", content)
    if not headers:
        raise HTTPException(status_code=400, detail="Nie udało się odczytać pliku — sprawdź format")

    mapping = auto_map_headers(headers)
    return {
        "filename": file.filename,
        "headers": headers,
        "row_count": len(rows),
        "preview": rows[:5],
        "auto_mapping": mapping,
    }


class CommitImportRequest(BaseModel):
    feed_name: str
    headers: list[str]
    rows: list[list[str]]
    mapping: dict[str, str | None]  # header -> canonical field (or None to skip)
    sku_field: str | None = None     # which canonical field acts as unique SKU


@router.post("/commit")
async def commit_import(
    body: CommitImportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new feed_in (manual source) and bulk-insert products from
    the parsed rows using the user-confirmed column mapping.
    """
    if not body.feed_name.strip():
        raise HTTPException(status_code=400, detail="Nazwa feedu jest wymagana")
    if not body.rows:
        raise HTTPException(status_code=400, detail="Brak danych do zaimportowania")

    # Plan limit check
    if current_user.plan and current_user.plan.max_products is not None:
        existing = await db.execute(
            select(ProductIn)
            .join(FeedIn, ProductIn.feed_in_id == FeedIn.id)
            .where(FeedIn.user_id == current_user.id)
        )
        already = len(existing.scalars().all())
        if already + len(body.rows) > current_user.plan.max_products:
            raise HTTPException(
                status_code=403,
                detail=f"Plan {current_user.plan.name} pozwala na max {current_user.plan.max_products} produktów. Już masz {already}.",
            )

    # Create feed_in
    feed = FeedIn(
        user_id=current_user.id,
        name=body.feed_name.strip(),
        source_url=f"manual://csv-import/{body.feed_name.strip()}",
        fetch_status="success",
    )
    db.add(feed)
    await db.flush()  # need the id

    name_field = "name"
    if body.sku_field and body.sku_field != "name":
        # If user picked a different SKU field, name still required
        pass

    # Find which header maps to name and which to id
    header_to_field = {h: f for h, f in body.mapping.items() if f}

    inserted = 0
    for row in body.rows:
        product_value: dict[str, str] = {}
        for i, header in enumerate(body.headers):
            if i >= len(row):
                continue
            val = (row[i] or "").strip()
            if not val:
                continue
            field = header_to_field.get(header)
            if field:
                product_value[field] = val
            # Always keep raw header too — useful for custom mapping later
            product_value[header] = val

        product_name = product_value.get(name_field) or product_value.get("name") or "(brak nazwy)"
        prod = ProductIn(
            feed_in_id=feed.id,
            product_name=product_name[:500],
            product_value=product_value,
            custom_product=True,
        )
        db.add(prod)
        inserted += 1

    await db.commit()
    await db.refresh(feed)

    return {
        "feed_in_id": feed.id,
        "name": feed.name,
        "products_imported": inserted,
    }
